"""Libro Excel del panel de administración: resumen, series y tablas de detalle."""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from Aplicacion.Servicios import certificado


def _border_thin() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_row(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="1E40AF")
    font = Font(bold=True, color="FFFFFF", size=11)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border_thin()


def _data_cell(cell, wrap: bool = False) -> None:
    cell.border = _border_thin()
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)


def _autosize_columns(ws, max_width: int = 55, min_width: float = 10.0) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = min_width
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            lines = str(v).splitlines()
            n = max((len(line) for line in lines), default=0)
            maxlen = max(maxlen, min(n, max_width))
        ws.column_dimensions[letter].width = min(max(maxlen + 2.2, min_width), max_width + 6)


def build_dashboard_excel_bytes() -> Optional[bytes]:
    try:
        return _build_dashboard_excel_bytes_impl()
    except Exception:
        return None


def _build_dashboard_excel_bytes_impl() -> bytes:
    insights: dict[str, Any] = certificado.obtener_dashboard_insights()
    wb = Workbook()
    ws = wb.active
    ws.title = "Variables por alumno"

    title_font = Font(bold=True, size=14, color="1E3A8A")
    sub_font = Font(size=10, color="666666")

    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.merge_cells("A1:I1")
    ws["A1"] = "Panel de certificados — variables por alumno"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Exportado: {now}"
    ws["A2"].font = sub_font

    headers = (
        "ID certificado",
        "Alumno",
        "N° validaciones",
        "Certificado validado",
        "Tiene TV",
        "TV (s)",
        "Tiene TGC",
        "TGC (s)",
        "Fuente",
    )
    header_row = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    _header_row(ws, header_row, len(headers))

    tv_rows = insights.get("tvByCertificate") or []
    tgc_rows = insights.get("genByCertificate") or []

    merged: dict[str, dict[str, Any]] = {}
    for r in tv_rows:
        cid = str(r.get("id") or "").strip()
        if not cid:
            continue
        merged[cid] = {
            "id": cid,
            "name": str(r.get("name") or "").strip(),
            "valid": bool(r.get("valid")),
            "hasTv": bool(r.get("hasTv")),
            "tv": float(r.get("tv") or 0.0),
            "validationCount": int(r.get("validationCount") or 0),
            "hasTgc": False,
            "tgc": 0.0,
            "source": "TV",
        }

    for r in tgc_rows:
        cid = str(r.get("id") or "").strip()
        if not cid:
            continue
        row = merged.setdefault(
            cid,
            {
                "id": cid,
                "name": str(r.get("name") or "").strip(),
                "valid": False,
                "hasTv": False,
                "tv": 0.0,
                "validationCount": 0,
                "hasTgc": False,
                "tgc": 0.0,
                "source": "TGC",
            },
        )
        row["name"] = row["name"] or str(r.get("name") or "").strip()
        row["hasTgc"] = bool(r.get("hasTgc"))
        row["tgc"] = float(r.get("tgc") or 0.0)
        row["source"] = "TV+TGC" if row["source"] == "TV" else "TGC"

    data_rows = sorted(
        merged.values(),
        key=lambda x: ((x.get("name") or "").lower(), (x.get("id") or "").lower()),
    )

    ri = header_row + 1
    for row in data_rows:
        ws.cell(row=ri, column=1, value=row["id"])
        ws.cell(row=ri, column=2, value=row["name"])
        ws.cell(row=ri, column=3, value=int(row.get("validationCount") or 0))
        ws.cell(row=ri, column=4, value="Sí" if row["valid"] else "No")
        ws.cell(
            row=ri,
            column=5,
            value="Sí" if row["hasTv"] else "No",
        )
        ws.cell(
            row=ri,
            column=6,
            value=round(float(row["tv"]), 6) if row["hasTv"] else "—",
        )
        ws.cell(row=ri, column=7, value="Sí" if row["hasTgc"] else "No")
        ws.cell(
            row=ri,
            column=8,
            value=round(float(row["tgc"]), 6) if row["hasTgc"] else "—",
        )
        ws.cell(row=ri, column=9, value=row["source"])
        for c in range(1, 10):
            _data_cell(ws.cell(row=ri, column=c), wrap=c in (2,))
        ri += 1

    _autosize_columns(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
