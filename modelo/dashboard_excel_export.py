"""Libro Excel del panel de administración: resumen, series y tablas de detalle."""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modelo import certificado


def _border_thin() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_row(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="1E40AF")
    font = Font(bold=True, color="FFFFFF", size=11)
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border_thin()


def _data_cell(cell, wrap: bool = False) -> None:
    cell.border = _border_thin()
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)


def _autosize_columns(ws, max_width: int = 55, min_width: float = 10.0) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        maxlen = min_width
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            lines = str(v).splitlines()
            n = max((len(line) for line in lines), default=0)
            maxlen = max(maxlen, min(n, max_width))
        ws.column_dimensions[letter].width = min(max(maxlen + 2.2, min_width), max_width + 6)


def build_dashboard_excel_bytes() -> Optional[bytes]:
    try:
        return _build_dashboard_excel_bytes_impl()
    except Exception:
        return None


def _build_dashboard_excel_bytes_impl() -> bytes:
    stats = certificado.obtener_estadisticas()
    insights: dict[str, Any] = certificado.obtener_dashboard_insights()

    wb = Workbook()
    title_font = Font(bold=True, size=14, color="1E3A8A")
    sub_font = Font(size=10, color="666666")

    # --- Hoja Resumen ---
    ws0 = wb.active
    ws0.title = "Resumen"
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws0.merge_cells("A1:B1")
    ws0["A1"] = "Panel de certificados — métricas"
    ws0["A1"].font = title_font
    ws0.merge_cells("A2:B2")
    ws0["A2"] = f"Exportado: {now}"
    ws0["A2"].font = sub_font

    status = insights.get("status") or {}
    total_certs = int(status.get("total") or 0)
    ver_total = int(stats.get("verCount") or 0)
    valid_n = int(stats.get("validCount") or 0)
    invalid_n = int(stats.get("invalidCount") or 0)
    rate = (100.0 * valid_n / ver_total) if ver_total > 0 else 0.0

    rows = [
        ("Certificados en base de datos", total_certs),
        ("— Activos", int(status.get("active") or 0)),
        ("— Revocados", int(status.get("revoked") or 0)),
        ("", ""),
        ("Total verificaciones (acumulado app)", ver_total),
        ("Verificaciones correctas", valid_n),
        ("Verificaciones incorrectas", invalid_n),
        ("Tasa de verificaciones correctas (%)", round(rate, 2)),
        ("", ""),
        ("Tiempo medio generación TGC (s)", stats.get("avgGenTime") or "0"),
        ("Tiempo medio verificación TV (s)", stats.get("avgVerTime") or "0"),
        ("Emisiones registradas (contador métricas)", int(stats.get("genCount") or 0)),
    ]
    start = 4
    ws0.cell(row=start, column=1, value="Indicador")
    ws0.cell(row=start, column=2, value="Valor")
    _header_row(ws0, start, 2)
    r = start + 1
    for label, val in rows:
        ws0.cell(row=r, column=1, value=label)
        ws0.cell(row=r, column=2, value=val)
        _data_cell(ws0.cell(row=r, column=1), wrap=True)
        _data_cell(ws0.cell(row=r, column=2))
        r += 1
    _autosize_columns(ws0)

    # --- Emisión y métricas mensuales ---
    ws1 = wb.create_sheet("Emisión mensual")
    headers = (
        "Período (mm/aaaa)",
        "Emitidos",
        "Activos",
        "Revocados",
        "TGC medio (s)",
        "TV medio (s)",
    )
    for i, h in enumerate(headers, 1):
        ws1.cell(row=1, column=i, value=h)
    _header_row(ws1, 1, len(headers))
    monthly = insights.get("monthly") or []
    for ri, m in enumerate(monthly, 2):
        ws1.cell(row=ri, column=1, value=m.get("label") or "")
        ws1.cell(row=ri, column=2, value=int(m.get("emitted") or 0))
        ws1.cell(row=ri, column=3, value=int(m.get("active") or 0))
        ws1.cell(row=ri, column=4, value=int(m.get("revoked") or 0))
        ws1.cell(row=ri, column=5, value=round(float(m.get("avgGen") or 0), 6))
        ws1.cell(row=ri, column=6, value=round(float(m.get("avgVer") or 0), 6))
        for c in range(1, 7):
            _data_cell(ws1.cell(row=ri, column=c))
    _autosize_columns(ws1)

    # --- TV por certificado ---
    ws2 = wb.create_sheet("TV por certificado")
    h2 = ("ID certificado", "Estudiante", "TV (s)", "Última verif. válida")
    for i, h in enumerate(h2, 1):
        ws2.cell(row=1, column=i, value=h)
    _header_row(ws2, 1, len(h2))
    for ri, row in enumerate(insights.get("tvByCertificate") or [], 2):
        ws2.cell(row=ri, column=1, value=row.get("id") or "")
        ws2.cell(row=ri, column=2, value=row.get("name") or "")
        tv = row.get("tv") if row.get("hasTv") else None
        ws2.cell(row=ri, column=3, value=round(float(tv), 6) if tv is not None else "—")
        ws2.cell(row=ri, column=4, value="Sí" if row.get("valid") else "No")
        for c in range(1, 5):
            _data_cell(ws2.cell(row=ri, column=c), wrap=c == 2)
    _autosize_columns(ws2)

    # --- TGC por certificado ---
    ws3 = wb.create_sheet("TGC por certificado")
    h3 = ("ID certificado", "Estudiante", "TGC (s)")
    for i, h in enumerate(h3, 1):
        ws3.cell(row=1, column=i, value=h)
    _header_row(ws3, 1, len(h3))
    for ri, row in enumerate(insights.get("genByCertificate") or [], 2):
        ws3.cell(row=ri, column=1, value=row.get("id") or "")
        ws3.cell(row=ri, column=2, value=row.get("name") or "")
        tgc = row.get("tgc") if row.get("hasTgc") else None
        ws3.cell(row=ri, column=3, value=round(float(tgc), 6) if tgc is not None else "—")
        for c in range(1, 4):
            _data_cell(ws3.cell(row=ri, column=c), wrap=c == 2)
    _autosize_columns(ws3)

    # --- Top cursos / tipos ---
    ws4 = wb.create_sheet("Top cursos y tipos")
    ws4.merge_cells("A1:B1")
    c = ws4["A1"]
    c.value = "Top 5 cursos (por cantidad de certificados)"
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(vertical="center")
    ws4.cell(row=2, column=1, value="Curso")
    ws4.cell(row=2, column=2, value="Cantidad")
    _header_row(ws4, 2, 2)
    r = 3
    for item in insights.get("topCourses") or []:
        ws4.cell(row=r, column=1, value=item.get("name") or "")
        ws4.cell(row=r, column=2, value=int(item.get("count") or 0))
        _data_cell(ws4.cell(row=r, column=1), wrap=True)
        _data_cell(ws4.cell(row=r, column=2))
        r += 1
    r += 1
    ws4.merge_cells(f"A{r}:B{r}")
    t = ws4.cell(row=r, column=1)
    t.value = "Top 5 tipos de credencial"
    t.font = Font(bold=True, size=12)
    t.alignment = Alignment(vertical="center")
    r += 1
    ws4.cell(row=r, column=1, value="Tipo")
    ws4.cell(row=r, column=2, value="Cantidad")
    _header_row(ws4, r, 2)
    r += 1
    for item in insights.get("topTypes") or []:
        ws4.cell(row=r, column=1, value=item.get("name") or "")
        ws4.cell(row=r, column=2, value=int(item.get("count") or 0))
        _data_cell(ws4.cell(row=r, column=1), wrap=True)
        _data_cell(ws4.cell(row=r, column=2))
        r += 1
    _autosize_columns(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
